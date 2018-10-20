#coding=utf-8   //这句是使用utf8编码方式方法， 可以单独加入python头使用
# from __future__ import print_function
import json  
from openpyxl import load_workbook
import yaml
import os
import shutil
import time
import logger

def gen_report_name(rootpath=".", reportname="report.html"):
    #定义date为日期，time为时间 
    date=time.strftime("%Y%m%d")
    # time=time.strftime("%Y%m%d%H%M%S")
  
    #定义path为文件路径，目录级别，可根据实际情况自定义修改
    path= rootpath + "/" + date
    fullpath = os.path.abspath(path)
  
    #定义报告文件路径和名字，路径为前面定义的path，名字为report（可自定义），格式为.html  
    report_path = path+"/"+reportname
      
    #判断是否定义的路径目录存在，不能存在则创建  
    if not os.path.exists(fullpath):
        os.makedirs(fullpath)

    return os.path.abspath(report_path)

def dump_dic(target, dic_list):
    f=open(target,"w", encoding="utf-8")
    for dic in dic_list:  
        json.dump(dic, f, ensure_ascii=False)
        f.write("\n")
    f.close()

def load_dic(src, dic_list):
    f=open(src,"r", encoding="utf-8")
    # data=list()
    for line in f:
        l=line.strip()
        if(len(l)==0 or l[0]=="#"):
            continue
        else:
            dic_list.append(json.loads(l))
    # logger.debug(dic_list)
    f.close()

def load_xls(src, sheetname=None):
    dic_list=list()
    wb = load_workbook(filename = src)
    # ws=wb.get_sheet_by_name(testcase_sheet)
    if(sheetname==None):
        ws=wb.get_active_sheet()
    else:
        ws=wb[sheetname]
    nr=ws.max_row
    nc=ws.max_column
    keys=list()
    data=list()
    for j in range(nc):
        keys.append(ws.cell(1,j+1).value)
    for i in range(nr-1):
        data.clear()
        for j in range(nc):
            data.append(ws.cell(i+2, j+1).value)
        dic_list.append(dict(zip(keys, data)))
    wb.close()
    return dic_list

def sub_var_value(src, var_dic):
    target=""
    target=src.format(**var_dic)
    return target

def load_from_yaml(src):
    f = open(src, encoding='utf-8')
    x = yaml.load(f)
    f.close()

    return x

# def setup_logging(log_cfg, logger_name="fileLogger"):
#     global logger

#     # log_cfg=load_from_yaml(path)
#     # if(loggername in log_cfg["loggers"]):
#     for loggername in log_cfg["loggers"]:
#         handlers=log_cfg["loggers"][loggername]["handlers"]
#         for handler in handlers:
#             if("filename" not in log_cfg["handlers"][handler]):
#                 continue
#             filename=log_cfg["handlers"][handler]["filename"]
#             head, tail=os.path.split(filename)
#             if head and tail and not os.path.exists(head):
#                 os.makedirs(head)
#             # os.makedirs(head, exist_ok=True)
#     # else:
#     #     print("Logger[%s] not int log configfile" %(loggername))

#     logging.config.dictConfig(log_cfg)
#     logger = logging.getLogger(logger_name)
#     return logger

# def load_config(yaml_file):
#     conf_dic=load_from_yaml(yaml_file)
#     logger=setup_logging(conf_dic["LOGGING"], logger_name=conf_dic["loggername"])
#     reportdir=conf_dic["reportdir"]
#     if not os.path.exists(reportdir):
#         os.makedirs(reportdir)
#     return conf_dic, logger

def find_config(dic_list, configid, key="ID"):
    for dic in dic_list:
        # if(dic.has_key(key)):
        if(key in dic and dic[key]==configid):
            return dic
        else:
            continue
    return None

def delete_file(filename):
    ret=0
    if(os.access(filename, os.R_OK)):
        try:
            os.remove(filename)
            # cmd="del \""+ orig_file + "\""
            # print(cmd)
            # ret=os.system(cmd)
        # except Exception,e:
        except OSError as e:
            logger.log_error("\ndelete file <%s> error, except[%s]\n" %(filename, e))
            ret=-1
        else:
            pass

    return ret

def change_file_name(src, target):
    ret=0
    i=0
    p_f_l=-1
    while(i<20):
        if(os.access(src, os.R_OK)):
            statinfo=os.stat(src)
            f_l=statinfo.st_size
            #文件至少有1个字节
            if(p_f_l==f_l and f_l>1):
                break
            else:
                p_f_l=f_l
                time.sleep(3)
        else:
            ret=-1
            time.sleep(3)
        i+=1
    if(i==20):
        logger.log_error("Original file[%s] is used or not exist" %(src))
        return -1
    else:
        ret=0
        try:
            # cmd="move \""+ orig_file + "\" \"" + target_file + "\""
            # print(cmd)
            # ret=os.system(cmd)
            head, tail = os.path.split(target)
            if head and tail and not os.path.exists(head):
                os.makedirs(head)
            # os.rename(src, target)
            shutil.move(src, target)
        except OSError as e:
            logger.log_error("\nrename file <%s> --> <%s> error except[%s]\n" %(src, target, e))
            return -2
        else:
            pass
    return ret

if __name__=="__main__":
    action_list=[
        {"NAME":"损益明细", "ACTION":"GET", "TAG":"", "CONTENT":"http://it.yusys.com.cn/yusys/DeptProfitAndLossDetail/toList.asp", "AFTER":"sleep(5)"},
        {"NAME":"核算类别", "FIND":"CSS", "ACTION":"CLICK", "TAG":"#s2id_accountingmethod > a:nth-child(1) > span:nth-child(1)"},
        {"NAME":"完工百分比法", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(2) > div:nth-child(1)[title='完工百分比法']"},
        {"NAME":"验收确认法", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(3) > div:nth-child(1)[title='验收确认法']"},

        {"NAME":"报表口径", "FIND":"CSS", "ACTION":"CLICK", "TAG":"#s2id_caliber > a:nth-child(1) > span:nth-child(1)"},
        {"NAME":"考核口径", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(2) > div:nth-child(1)[title='考核口径']"},
        {"NAME":"管理口径", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(3) > div:nth-child(1)[title='管理口径']"},

        {"NAME":"查询时间", "FIND":"ID", "ACTION":"SEND_KEYS", "TAG":"searchtime", "CONTENT":"$Y_M"},

        {"NAME":"项目分类", "FIND":"CSS", "ACTION":"CLICK", "TAG":"#s2id_projclassify > a:nth-child(1) > span:nth-child(1)"},
        {"NAME":"项目", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(2) > div:nth-child(1)[title='项目']"},
        {"NAME":"非项目", "FIND":"CSS", "ACTION":"CLICK", "TAG":"li.select2-results-dept-0:nth-child(3) > div:nth-child(1)[title='非项目']"},

        {"NAME":"查询", "FIND":"ID", "ACTION":"CLICK", "TAG":"query"},

        {"NAME":"报表结果", "FIND":"TAG", "ACTION":"SWITCH_FRAME", "TAG":"iframe", "CONTENT":"runqian"},
        {"NAME":"另存excel", "FIND":"CSS", "ACTION":"CLICK", "TAG":"#functionBar > a:nth-child(1) > img:nth-child(1)[alt='存为Excel']", "BEFORE":"del_file", "AFTER":"rename_file"}
    ]
    # file_name="test-sy.json"
    # dump_dic(file_name, action_list)

    file_name="F:/workspace/python/excel/report-conf.json"
    new_action_list=list()
    load_dic(file_name, new_action_list)
    print(new_action_list)
    f=open("X:/TEMP/y.yaml", "w", encoding="utf-8")
    yaml.dump(new_action_list, stream=f, allow_unicode=True, encoding="utf-8")
    f.close()

    # # change_file_name("test-sy.json", "test-sy.json")
    # change_file_name("y.json", "test-sy.json")

    # load_xls("损益-测试案例.xlsx", "损益", "损益.DATA", testcase_list, testdata_list)

    # testcase_list=list()
    # load_xls("损益-测试案例.xlsx", "损益", testcase_list)
    # for i, data in enumerate(testcase_list):
    #     print("STEP[%d][%s]" %(i,data))

    # testdata_list=list()
    # load_xls("损益-测试案例.xlsx", "损益.DATA", testdata_list)
    # for i, data in enumerate(testdata_list):
    #     print("TESTDATA[%d][%s]" %(i,data))

    # src="li.select2-results-dept-0 > div.select2-result-label[title='$核算类别']"
    # # src="li.select2-results-dept-0 > div.select2-result-label[title='$核算别']"
    # target=sub_var_value(src, testdata_list[2])
    # print("target[%s]" %(target))

    # conf_dic, logger=load_config("F:/workspace/python/unittest-yusys/yusys.yaml")
    # print(type(conf_dic))
    # print(conf_dic)
    # # logger=setup_logging(conf_dic["log"])
    # logger.info('foorbar')
    # logger.error('this is test error')